# orchestrator/scriptlets/python/dash_live_mem.py
from typing import List
import io, json, threading, time, socket
from datetime import datetime

from dash import Dash, dcc, html, Input, Output, State, MATCH, ALL, no_update
import plotly.graph_objs as go
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series

def _is_port_free(port: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.25)
        return s.connect_ex(("127.0.0.1", port)) != 0

def _choose_port(preferred: int, fallbacks: List[int]):
    if _is_port_free(preferred):
        return preferred
    for p in fallbacks:
        if _is_port_free(p):
            return p
    return None

def _get_headers(ctx) -> List[str]:
    cfg = ctx.get("stream_config") or {}
    return ["Timestamp"] + list(cfg.get("headers") or [])

def _get_rows_slice(ctx, last_n: int):
    rows = ctx.get("stream_rows") or []
    if last_n and last_n > 0:
        return rows[-last_n:]
    return rows

def _make_fig(rows, y_cols):
    fig = go.Figure()
    if not rows:
        return fig
    x = [r.get("Timestamp") for r in rows]
    for col in (y_cols or []):
        if col == "Timestamp":
            continue
        ys = []
        for r in rows:
            v = r.get(col, None)
            try:
                ys.append(float(v))
            except Exception:
                ys.append(None)
        fig.add_trace(go.Scatter(x=x, y=ys, mode="lines", name=col))
    fig.update_layout(
        hovermode="x unified",
        margin=dict(l=40, r=20, t=40, b=40),
        xaxis_title="Time",
        yaxis_title="Value"
    )
    return fig

def _build_excel_bytes(rows, y_cols, title: str):
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"
    cols = ["Timestamp"] + [c for c in (y_cols or []) if c != "Timestamp"]
    ws_data.append(cols)
    for r in rows:
        ws_data.append([r.get(c) for c in cols])

    # Chart sheet
    if ws_data.max_row >= 2 and len(cols) >= 2:
        ws_chart = wb.create_sheet("Charts")
        lc = LineChart()
        lc.title = f"{title} export"
        lc.x_axis.title = "Time"
        lc.y_axis.title = "Value"
        max_row = ws_data.max_row
        xref = Reference(ws_data, min_col=1, min_row=2, max_row=max_row)
        for j in range(2, len(cols) + 1):
            yref = Reference(ws_data, min_col=j, min_row=1, max_row=max_row)
            lc.append(Series(yref, xref, title=cols[j - 1]))
        ws_chart.add_chart(lc, "A1")

    ws_notes = wb.create_sheet("note")
    ws_notes.append(["timestamp", "tester_note"])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

def _ctx_dump_loop(ctx, period_sec: int, title: str):
    # Dump CTX summary to stdout periodically (visible in step logs)
    while ctx.get("_dash_running", False):
        try:
            keys = list(ctx.keys())
            sample_rows = (_get_rows_slice(ctx, 3) or [])  # only top 3 recent rows
            print(f"[CTX DUMP] {datetime.utcnow().isoformat()}Z {title}")
            print(json.dumps({
                "ctx_keys": keys,
                "rows_sample_count": len(sample_rows),
                "rows_last_sample": (sample_rows[-1] if sample_rows else None)
            }, indent=2), flush=True)
        except Exception as e:
            print(f"[CTX DUMP ERROR] {e}", flush=True)
        time.sleep(period_sec)

def _chart_block(chart_id: int, headers: List[str], default_y_cols: List[str], draw_last_n: int):
    options = [{"label": h, "value": h} for h in headers if h != "Timestamp"]
    return html.Div(
        id={"type": "chart-block", "index": chart_id},
        children=[
            html.Div([
                dcc.Dropdown(
                    id={"type": "ycols", "index": chart_id},
                    options=options,
                    value=[c for c in default_y_cols if c in headers],
                    multi=True,
                    placeholder="Select columns to plot"
                ),
                dcc.Slider(
                    id={"type": "lastn", "index": chart_id},
                    min=100, max=20000, step=100,
                    value=draw_last_n
                ),
            ], style={"display":"grid","gridTemplateColumns":"1fr 1fr","gap":"8px","margin":"8px 0"}),
            dcc.Graph(id={"type":"live-graph", "index": chart_id})
        ],
        style={"border":"1px solid #ddd","borderRadius":"6px","padding":"8px","marginBottom":"10px"}
    )

def run(ctx, params):
    """
    Launch Dash app that:
      - Plots ctx['stream_rows'] in one or more charts
      - Lets user Add/Remove chart blocks with their own selectors
      - Exports the FIRST chart's selection to Excel via button
      - Logs a CTX snapshot every ctx_dump_sec seconds
    params:
      host, port, fallback_ports, default_y_cols, title, poll_ms, draw_last_n, ctx_dump_sec
    """
    host = params.get("host", "0.0.0.0")
    port = int(params.get("port", 8050))
    fallback_ports = params.get("fallback_ports") or [8051, 8052, 8060, 8080]
    default_y_cols = params.get("default_y_cols") or []
    title = params.get("title", "Thermal Stream (Shared Memory)")
    poll_ms = int(params.get("poll_ms", 1000))
    draw_last_n = int(params.get("draw_last_n", 5000))
    ctx_dump_sec = int(params.get("ctx_dump_sec", 10))

    chosen_port = _choose_port(port, fallback_ports)
    if chosen_port is None:
        raise RuntimeError("No free port available for Dash (preferred and fallbacks taken).")
    print(f"[Dash] Using port {chosen_port}", flush=True)

    app = Dash(__name__)
    headers = _get_headers(ctx)

    # Initial single chart
    initial_chart_id = 1

    app.layout = html.Div([
        html.H3(title, style={"margin":"12px 0"}),
        html.Div([
            html.Button("Add chart", id="add-chart", n_clicks=0),
            html.Button("Remove last chart", id="remove-chart", n_clicks=0),
            html.Button("Export FIRST chart to Excel", id="export-btn", n_clicks=0, style={"marginLeft":"auto"})
        ], style={"display":"flex","gap":"8px","alignItems":"center","margin":"8px 0"}),

        # Container for dynamic chart blocks
        html.Div(id="charts-container"),

        # Global polling interval shared by all charts
        dcc.Interval(id="poll", interval=poll_ms, n_intervals=0),

        # Downloader for export
        dcc.Download(id="download-excel"),

        # Store the list of chart IDs
        dcc.Store(id="chart-ids", data=[initial_chart_id]),

        # Store headers (so we can rebuild dropdown options if needed)
        dcc.Store(id="headers-store", data=headers),

        # Defaults for new charts
        dcc.Store(id="defaults-store", data={
            "default_y_cols": default_y_cols,
            "draw_last_n": draw_last_n
        })
    ], style={"maxWidth":"1100px","margin":"0 auto","padding":"8px"})

    # Initialize the charts container once with initial block
    @app.callback(
        Output("charts-container", "children"),
        Input("chart-ids", "data"),
        State("headers-store", "data"),
        State("defaults-store", "data")
    )
    def _render_chart_blocks(chart_ids, headers_state, defaults):
        if not chart_ids:
            return []
        blocks = []
        for cid in chart_ids:
            blocks.append(_chart_block(
                chart_id=cid,
                headers=headers_state or headers,
                default_y_cols=defaults.get("default_y_cols") or [],
                draw_last_n=int(defaults.get("draw_last_n") or 5000)
            ))
        return blocks

    # Add/Remove chart IDs
    @app.callback(
        Output("chart-ids", "data"),
        Input("add-chart", "n_clicks"),
        Input("remove-chart", "n_clicks"),
        State("chart-ids", "data"),
        prevent_initial_call=True
    )
    def _mutate_charts(add_clicks, remove_clicks, ids):
        # Determine which button fired last
        ctx_trig = dash.callback_context.triggered[0]["prop_id"].split(".") if dash.callback_context.triggered else ""
        ids = list(ids or [])
        if ctx_trig == "add-chart":
            next_id = (max(ids) + 1) if ids else 1
            ids.append(next_id)
        elif ctx_trig == "remove-chart":
            if ids:
                ids.pop()
        return ids

    # Update each chart figure independently (pattern-matching callback)
    @app.callback(
        Output({"type":"live-graph", "index": MATCH}, "figure"),
        Input("poll", "n_intervals"),
        State({"type":"ycols", "index": MATCH}, "value"),
        State({"type":"lastn", "index": MATCH}, "value"),
        prevent_initial_call=False
    )
    def _refresh_chart(_n, ycols, lastn):
        rows = _get_rows_slice(ctx, int(lastn or draw_last_n))
        return _make_fig(rows, ycols or [])

    # Export: use the FIRST chart's selectors, if present
    @app.callback(
        Output("download-excel", "data"),
        Input("export-btn", "n_clicks"),
        State("chart-ids", "data"),
        State({"type":"ycols", "index": ALL}, "value"),
        State({"type":"lastn", "index": ALL}, "value"),
        prevent_initial_call=True
    )
    def _export_first(n_clicks, ids, ycols_all, lastn_all):
        if not ids:
            return no_update
        # Map ycols_all/lastn_all to ids by position since the Store/render preserve order
        first_idx = 0
        ycols_first = (ycols_all[first_idx] if ycols_all and len(ycols_all) > 0 else None) or []
        lastn_first = int((lastn_all[first_idx] if lastn_all and len(lastn_all) > 0 else draw_last_n) or draw_last_n)
        rows = _get_rows_slice(ctx, lastn_first)
        if not rows or not ycols_first:
            return no_update
        xlsx = _build_excel_bytes(rows, ycols_first, title)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return dcc.send_bytes(xlsx, filename=f"chart_export_{stamp}.xlsx")

    # Start the periodic CTX dump thread
    ctx["_dash_running"] = True
    t = threading.Thread(target=_ctx_dump_loop, args=(ctx, ctx_dump_sec, title), daemon=True)
    t.start()

    try:
        app.run_server(host=host, port=chosen_port, debug=False, use_reloader=False)
    finally:
        ctx["_dash_running"] = False

    return {"ok": True, "host": host, "port": chosen_port}
