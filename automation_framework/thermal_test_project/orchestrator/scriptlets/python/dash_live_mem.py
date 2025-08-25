# orchestrator/scriptlets/python/dash_live_mem.py
from typing import List
import io
from datetime import datetime

import dash
from dash import Dash, dcc, html, Input, Output, State, no_update
import dash_bootstrap_components as dbc
import plotly.graph_objs as go

# Excel export
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series

def _get_headers(ctx) -> List[str]:
    cfg = ctx.get("stream_config") or {}
    return ["Timestamp"] + list(cfg.get("headers") or [])

def _get_rows_slice(ctx, last_n: int):
    rows = ctx.get("stream_rows") or []
    if last_n and last_n > 0:
        return rows[-last_n:]
    return rows

def _make_fig(rows, y_cols):
    if not rows:
        return go.Figure()
    x = [r.get("Timestamp") for r in rows]
    fig = go.Figure()
    for col in y_cols or []:
        if col == "Timestamp":
            continue
        y = []
        for r in rows:
            v = r.get(col, None)
            try:
                y.append(float(v))
            except Exception:
                y.append(None)
        fig.add_trace(go.Scatter(x=x, y=y, mode="lines", name=col))
    fig.update_layout(
        hovermode="x unified",
        margin=dict(l=40, r=20, t=40, b=40),
        xaxis_title="Time",
        yaxis_title="Value"
    )
    return fig

def _build_excel_bytes(rows, y_cols, title: str):
    """
    Build an in-memory Excel with:
    - Data sheet: Timestamp + selected y columns
    - Charts sheet: line chart mirroring those columns
    - note sheet: placeholder
    """
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"

    cols = ["Timestamp"] + [c for c in (y_cols or []) if c != "Timestamp"]
    ws_data.append(cols)

    for r in rows:
        ws_data.append([r.get(c) for c in cols])

    # Charts
    ws_chart = wb.create_sheet("Charts")
    if ws_data.max_row >= 2 and len(cols) >= 2:
        lc = LineChart()
        lc.title = f"{title} export"
        lc.x_axis.title = "Time"
        lc.y_axis.title = "Value"
        max_row = ws_data.max_row

        # X axis from Data!A2:A{max}
        xref = Reference(ws_data, min_col=1, min_row=2, max_row=max_row)
        # Each Y series from Data!B1:B{max}, C1:C{max}, ...
        for j in range(2, len(cols) + 1):
            yref = Reference(ws_data, min_col=j, min_row=1, max_row=max_row)
            ser = Series(yref, xref, title=cols[j - 1])
            lc.append(ser)
        ws_chart.add_chart(lc, "A1")

    # note
    ws_notes = wb.create_sheet("note")
    ws_notes.append(["timestamp", "tester_note"])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

def run(ctx, params):
    """
    Orchestrator entrypoint. Launches a Dash app that reads live data
    from ctx["stream_rows"] and renders a line chart. Includes an
    'Export to Excel' button to download the currently visible data.
    """
    host = params.get("host", "0.0.0.0")
    port = int(params.get("port", 8052))
    default_y_cols = params.get("default_y_cols") or []
    title = params.get("title", "Live Stream (Shared Memory)")
    poll_ms = int(params.get("poll_ms", 1000))
    draw_last_n = int(params.get("draw_last_n", 5000))

    app = Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])
    headers = _get_headers(ctx)
    y_options = [{"label": h, "value": h} for h in headers if h != "Timestamp"]

    app.layout = dbc.Container([
        html.H3(title, className="mt-3"),
        dbc.Row([
            dbc.Col(dcc.Dropdown(
                id="ycols",
                options=y_options,
                value=[c for c in default_y_cols if c in headers],
                multi=True,
                placeholder="Select columns to plot"
            ), md=6),
            dbc.Col(dcc.Slider(
                id="lastn",
                min=100, max=20000, step=100,
                value=draw_last_n,
                marks=None,
                tooltip={"placement": "bottom", "always_visible": False}
            ), md=4),
            dbc.Col(dbc.Button("Export to Excel", id="export-btn", color="primary", className="w-100"), md=2),
        ], className="mb-2"),
        dcc.Graph(id="live-graph"),
        dcc.Interval(id="poll", interval=poll_ms, n_intervals=0),
        dcc.Download(id="download-excel")
    ], fluid=True)

    @app.callback(
        Output("live-graph", "figure"),
        Input("poll", "n_intervals"),
        State("ycols", "value"),
        State("lastn", "value")
    )
    def refresh(_n, ycols, lastn):
        rows = _get_rows_slice(ctx, int(lastn or draw_last_n))
        return _make_fig(rows, ycols or [])

    @app.callback(
        Output("download-excel", "data"),
        Input("export-btn", "n_clicks"),
        State("ycols", "value"),
        State("lastn", "value"),
        prevent_initial_call=True
    )
    def do_export(n_clicks, ycols, lastn):
        rows = _get_rows_slice(ctx, int(lastn or draw_last_n))
        if not rows or not (ycols or []):
            return no_update
        try:
            xlsx_bytes = _build_excel_bytes(rows, ycols or [], title)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            return dcc.send_bytes(xlsx_bytes, filename=f"chart_export_{stamp}.xlsx")
        except Exception as e:
            print(f"[Export Error] {e}", flush=True)
            return no_update

    # Blocking call; orchestrator will keep this step 'running' until stopped.
    app.run_server(host=host, port=8053, debug=False)
    return {"ok": True, "host": host, "port": port}
