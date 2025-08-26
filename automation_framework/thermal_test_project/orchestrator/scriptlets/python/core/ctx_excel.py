"""
ctx_excel.py
------------
Writes ctx keys as sheets in an Excel file, with optional charts.
Write multiple ctx keys to multi-sheet Excel, with chart support.

Usage:
    run(ctx, {"sheets": {"foo": "foo", "bar": "bar"}, "out": "Data/all.xlsx", "charts": [{"sheet": "foo", "x": "col1", "y": ["col2"]}]})

Description:
    - Uses openpyxl.
"""
import openpyxl
from openpyxl.chart import LineChart, Reference

def run(ctx, params):
    wb = openpyxl.Workbook()
    sheets = params["sheets"]
    for key, sheet_name in sheets.items():
        ws = wb.create_sheet(sheet_name)
        data = ctx[key]
        for row in data:
            ws.append(row)
    for chart in params.get("charts", []):
        ws = wb[chart["sheet"]]
        x_col = chart["x"]
        y_cols = chart["y"]
        x_idx = ws[1].index(x_col) + 1
        for y in y_cols:
            y_idx = ws[1].index(y) + 1
            chart_obj = LineChart()
            chart_obj.add_data(Reference(ws, min_col=y_idx, min_row=2, max_row=ws.max_row), titles_from_data=False)
            chart_obj.set_categories(Reference(ws, min_col=x_idx, min_row=2, max_row=ws.max_row))
            ws.add_chart(chart_obj, f"E{ws.max_row+2}")
    wb.save(params["out"])
    return {"written": params["out"]}