"""
ctx_excel_advanced.py
---------------------
Writes ctx keys as sheets in an Excel file, with optional charts and formatting.
Write multiple ctx keys to multi-sheet Excel, with support for charts and formatting.

Usage:
    run(ctx, {
        "sheets": {"foo": "FooSheet", "bar": "BarSheet"},
        "out": "Data/all.xlsx",
        "charts": [
            {"sheet": "FooSheet", "x": "Time", "y": ["Value1", "Value2"], "type": "line"}
        ]
    })

Description:
    - Uses openpyxl for Excel writing and charting.
    - Supports multi-sheet, column headers, and line/bar charts.
"""
import openpyxl
from openpyxl.chart import LineChart, BarChart, Reference

def run(ctx, params):
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    sheets = params["sheets"]
    for key, sheet_name in sheets.items():
        ws = wb.create_sheet(sheet_name)
        data = ctx[key]
        # Write header if data is list of dicts
        if isinstance(data, list) and data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h, "") for h in headers])
        elif isinstance(data, list):
            for row in data:
                ws.append(row)
    for chart in params.get("charts", []):
        ws = wb[chart["sheet"]]
        headers = [cell.value for cell in ws[1]]
        x_idx = headers.index(chart["x"]) + 1
        for y in chart["y"]:
            y_idx = headers.index(y) + 1
            if chart.get("type", "line") == "line":
                chart_obj = LineChart()
            else:
                chart_obj = BarChart()
            chart_obj.title = f"{y} vs {chart['x']}"
            chart_obj.add_data(Reference(ws, min_col=y_idx, min_row=1, max_row=ws.max_row), titles_from_data=True)
            chart_obj.set_categories(Reference(ws, min_col=x_idx, min_row=2, max_row=ws.max_row))
            ws.add_chart(chart_obj, f"E{ws.max_row+2}")
    wb.save(params["out"])
    return {"written": params["out"]}