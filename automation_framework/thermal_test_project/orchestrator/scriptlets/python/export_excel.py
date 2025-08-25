# Excel export from CSV and chart spec
import json, csv, os
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series

def run(params):
    # Extract parameters
    input_csv = params["input_csv"]
    chart_spec_json = params["chart_spec_json"]
    excel_out = params["excel_out"]

    # Load the chart spec JSON
    spec = json.load(open(chart_spec_json, "r"))

    # Create a new Excel workbook and prepare Data sheet
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"

    # Fill the Data sheet with CSV content
    with open(input_csv, "r", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            ws_data.append(row)

    # Extract headers from first row of Data sheet
    headers = [c.value for c in ws_data[1] if c.value]

    # Iterate through defined sheets and charts
    for sheet in spec.get("sheets", []):
        ws = wb.create_sheet(title=sheet["name"])
        # Place charts vertically with spacing
        for idx, ch in enumerate(sheet.get("charts", [])):
            # Only support line charts in this demo
            lc = LineChart()
            lc.title = ch.get("title", "")
            lc.y_axis.title = ch.get("y_axis_title", "")
            lc.x_axis.title = ch.get("x_axis_title", "")
            # Locate the x axis column index
            x_idx = headers.index(ch["x_col"]) + 1
            # Determine the last data row
            max_row = ws_data.max_row
            # Build series for each y column
            for ys in ch.get("y_series", []):
                # Find y column index
                y_idx = headers.index(ys["col"]) + 1
                # X values: from row 2 to max (skip header)
                xref = Reference(ws_data, min_col=x_idx, min_row=2, max_row=max_row)
                # Y values: include header for series title, from row 1 (title) to max
                yref = Reference(ws_data, min_col=y_idx, min_row=1, max_row=max_row)
                # Create the series; OpenPyXL uses header cell as title if provided
                ser = Series(yref, xref, title=ys.get("name"))
                # Append series to the chart
                lc.append(ser)
            # Position chart with some vertical spacing
            anchor_row = 2 + idx * 18
            ws.add_chart(lc, f"A{anchor_row}")

    # Add a "note" sheet with the expected header for dashboard annotations
    ws_notes = wb.create_sheet(title="note")
    ws_notes.append(["timestamp", "tester_note"])

    # Ensure output directory exists
    os.makedirs(os.path.dirname(excel_out), exist_ok=True)
    # Save workbook
    wb.save(excel_out)

    # Return path for logging
    return {"ok": True, "excel": excel_out}
